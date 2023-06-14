"""
    File created by Emili Zubillaga
    VenturiThrust Project 2017
"""


#==================
##### IMPORTS #####
#==================


from gui.thrust import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import sys
import os
from SerialCom import *
from PyQt5.QtCore import QThread, SIGNAL
from openpyxl import Workbook
import math

# Icon sizes definitions.
smallIconSize = 16
mediumIconSize = 32
bigIconSize = 64

# Wind speed variables definition
ORD = -36433.58251
X1 = 57232.6131
X2 = -37226.33966
X3 = 12830.17438
X4 = -2469.93537
X5 = 251.79544
X6 = -10.62012


# Functions and variables for executable generation.
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


executableGeneration = False
# executableGeneration = True
softwareVersionMajor = 0
softwareVersionMinor = 2
softwareVersionPatch = 0

developers = ["Emili Zubillaga"]


# ======================
# #### MAIN WINDOW #####
# ======================
class MainWindow(QMainWindow, Ui_ThrustStandMW):
    def __init__(self, parent = None):
        super(MainWindow,self).__init__(parent)
        QMainWindow.__init__(self)
        self.setupUi(self)

        self.com = SerialCom()
        self.myThread = traceThread()
        self.connect(self.myThread, SIGNAL("Serialtrace(QString)"), self.printTrace)
        self.printEnabled = False
        self.selectedPort = []
        self.rows = 1

        self.availableTasks = [
            'CALIBRATION',
            'THRUST',
            'WIND_SPEED',
            'MOTOR_SPEED',
            'POWER',
            'ALL_TASKS'
        ]

        """
            Menu Bar Configuration.
        """
        self.calibrationAction = QAction("ESC Calibration", self)
        self.exitAction = QAction("Exit", self)
        self.helpAction = QAction("Help...", self)
        self.versionAction = QAction("Version", self)
        self.mainMenu = self.menuBar()

        # Configuring the Program menu.
        fileMenu = self.mainMenu.addMenu("&File")
        connectionsMenu = fileMenu.addMenu("&Connections")
        index = 0
        for self.ports in list(self.com.availablePorts()):
            self.connectionAction = QAction(self)
            self.connectionAction.setText(self.ports[1])
            self.connectionAction.triggered.connect(lambda checked, index = self.ports: self.connectNewInterface(index))
            connectionsMenu.addAction(self.connectionAction)
            index = index + 1

        fileMenu.addAction(self.calibrationAction)
        fileMenu.addSeparator()
        fileMenu.addAction(self.exitAction)

        # Configuring the Program menu.
        helpMenu = self.mainMenu.addMenu("&About")
        helpMenu.addAction(self.helpAction)
        helpMenu.addAction(self.versionAction)

        # Set various default values.
        self.lbl_currentSpeedValueVar.setText("0")
        self.lbl_finalSpeedValueVar.setText("0")
        self.lbl_portConnected.setText("Not connected yet.")
        self.speedKnob.setNotchesVisible(True)
        self.txt_log.setStyleSheet("background-color: rgb(0, 0, 0);")
        self.txt_log.setTextColor(QColor(255, 255, 255))
        self.pb_stopLog.setEnabled(False)
        self.pb_stopManualControl.setEnabled(False)
        self.manualControlOnGoing = False
        self.pb_stopAutoControl.setEnabled(False)
        self.autoControlOnGoing = False

        """
            Set icons and icons' sizes for the push buttons (default).
            arrow-rewind-icon.png
        """
        if executableGeneration:
            self.pb_lessSpeed.setIcon(QIcon(resource_path("arrow-rewind-icon.png")))
            self.pb_moreSpeed.setIcon(QIcon(resource_path("arrow-forward-icon.png")))
            self.pb_setSpeedValue.setIcon(QIcon(resource_path("arrow-next-3-icon.png")))
            self.pb_stopManualControl.setIcon(QIcon(resource_path("Button-2-stop-icon.png")))
            self.pb_startManualControl.setIcon(QIcon(resource_path("Button-1-play-icon.png")))
            self.pb_timeElapsed.setIcon(QIcon(resource_path("arrow-next-3-icon.png")))
            self.pb_stopAutoControl.setIcon(QIcon(resource_path("Button-2-stop-icon.png")))
            self.pb_startAutoControl.setIcon(QIcon(resource_path("Button-1-play-icon.png")))
            self.pb_stopLog.setIcon(QIcon(resource_path("Button-2-stop-icon.png")))
            self.pb_startLog.setIcon(QIcon(resource_path("Button-1-play-icon.png")))
            self.pb_clearLog.setIcon(QIcon(resource_path("recyclebin-icon.png")))
            self.pb_restartBoard.setIcon(QIcon(resource_path("reload-icon.png")))
            self.pb_addCommand.setIcon(QIcon(resource_path("arrow-next-3-icon.png")))
            self.pb_eraseCommand.setIcon(QIcon(resource_path("arrow-back-icon.png")))
        else:
            self.pb_lessSpeed.setIcon(QIcon(os.path.join("img", "arrow-rewind-icon.png")))
            self.pb_moreSpeed.setIcon(QIcon(os.path.join("img", "arrow-forward-icon.png")))
            self.pb_setSpeedValue.setIcon(QIcon(os.path.join("img", "arrow-next-3-icon.png")))
            self.pb_stopManualControl.setIcon(QIcon(os.path.join("img", "Button-2-stop-icon.png")))
            self.pb_startManualControl.setIcon(QIcon(os.path.join("img", "Button-1-play-icon.png")))
            self.pb_timeElapsed.setIcon(QIcon(os.path.join("img", "arrow-next-3-icon.png")))
            self.pb_stopAutoControl.setIcon(QIcon(os.path.join("img", "Button-2-stop-icon.png")))
            self.pb_startAutoControl.setIcon(QIcon(os.path.join("img", "Button-1-play-icon.png")))
            self.pb_stopLog.setIcon(QIcon(os.path.join("img", "Button-2-stop-icon.png")))
            self.pb_startLog.setIcon(QIcon(os.path.join("img", "Button-1-play-icon.png")))
            self.pb_clearLog.setIcon(QIcon(os.path.join("img", "recyclebin-icon.png")))
            self.pb_restartBoard.setIcon(QIcon(os.path.join("img", "reload-icon.png")))
            self.pb_addCommand.setIcon(QIcon(os.path.join("img", "arrow-next-3-icon.png")))
            self.pb_eraseCommand.setIcon(QIcon(os.path.join("img", "arrow-back-icon.png")))

        self.pb_lessSpeed.setIconSize(QSize(smallIconSize, smallIconSize))
        self.pb_moreSpeed.setIconSize(QSize(smallIconSize, smallIconSize))
        self.pb_setSpeedValue.setIconSize(QSize(smallIconSize, smallIconSize))
        self.pb_stopManualControl.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_startManualControl.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_timeElapsed.setIconSize(QSize(smallIconSize, smallIconSize))
        self.pb_stopAutoControl.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_startAutoControl.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_stopLog.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_startLog.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_clearLog.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_restartBoard.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_addCommand.setIconSize(QSize(mediumIconSize, mediumIconSize))
        self.pb_eraseCommand.setIconSize(QSize(mediumIconSize, mediumIconSize))

        """
            Set default configuration for radio buttons.
        """
        # Create groups for each pair of radio buttons.
        self.rb_manualAutoGroup = QButtonGroup()
        self.rb_defaultAccGroup = QButtonGroup()

        # Arrange pairs of buttons into each group.
        self.rb_manualAutoGroup.addButton(self.rb_manualControl)
        self.rb_manualAutoGroup.addButton(self.rb_autoControl)
        self.rb_defaultAccGroup.addButton(self.rb_accDefault)
        self.rb_defaultAccGroup.addButton(self.rb_accEnabled)

        # Set default values for each pair of radio buttons
        self.rb_manualControl.setChecked(True)
        self.rb_accDefault.setChecked(True)

        # Disable elements that are not selected by the radio button.
        self.switchAutoManual()
        self.disableAcc()

        # Fill task list.
        self.list_availableCommands.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.list_shownCommands.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.fillAvailableTasksList()

        """
            Signals' connections configuration.
        """
        self.rb_manualControl.clicked.connect(self.switchAutoManual)
        self.rb_autoControl.clicked.connect(self.switchAutoManual)
        self.rb_accDefault.clicked.connect(self.disableAcc)
        self.rb_accEnabled.clicked.connect(self.disableAcc)
        self.pb_moreSpeed.clicked.connect(self.increaseSpeedValue)
        self.pb_lessSpeed.clicked.connect(self.decreaseSpeedValue)
        self.pb_setSpeedValue.clicked.connect(self.setSpeedValueManual)
        self.speedKnob.valueChanged.connect(self.manualControlKnob)
        self.txt_setSpeedValue.returnPressed.connect(self.setSpeedValueManual)
        self.sldr_finalSpeedValue.valueChanged.connect(self.autoControlSldr)
        self.pb_addCommand.clicked.connect(self.showSelectedTasks)
        self.pb_eraseCommand.clicked.connect(self.eraseSelectedTasks)
        self.pb_clearLog.clicked.connect(self.txt_log.clear)
        self.pb_startLog.clicked.connect(self.enablePrintTraces)
        self.pb_stopLog.clicked.connect(self.enablePrintTraces)
        self.cb_logEnable.stateChanged.connect(self.loggingEnable)
        self.pb_restartBoard.clicked.connect(self.resetCommunication)
        self.pb_startManualControl.clicked.connect(self.activateControl)
        self.pb_startAutoControl.clicked.connect(self.activateControl)
        self.pb_stopManualControl.clicked.connect(self.activateControl)
        self.pb_stopAutoControl.clicked.connect(self.activateControl)
        self.sb_acc.valueChanged.connect(self.setAcc)
        self.exitAction.triggered.connect(sys.exit)
        self.versionAction.triggered.connect(self.versionInfo)
        self.calibrationAction.triggered.connect(self.calibrationCommand)

    def fillAvailableTasksList(self):
        self.list_availableCommands.clear()
        for i in self.availableTasks:
            if i is not 'CALIBRATION' and i is not 'ALL_TASKS':
                self.list_availableCommands.addItem(i)

    def connectNewInterface(self, selectedPort):
        self.selectedPort = selectedPort
        self.com.open(self.selectedPort[0], 115200, 1)
        if self.com.isOpen():
            self.lbl_portConnected.setText(self.selectedPort[1])
            self.logMessage("INFO", "Successfully connected to " + self.selectedPort[1])
        self.myThread.start()

    def resetCommunication(self):
        try:
            self.com.close()
            self.connectNewInterface(self.selectedPort)
            if ser.isOpen():
                self.list_shownCommands.clear()
                self.fillAvailableTasksList()
            QMessageBox.information(self, "Reset successful",
                                    "The board has been successfully restarted!",
                                    QMessageBox.Ok)
        except:
            QMessageBox.warning(self, "Reset failed",
                                "Unable to reset.\nCheck if the board is actually connected or if the port is closed.",
                                QMessageBox.Ok)

    def enablePrintTraces(self):
        if self.com.isOpen():
            if self.printEnabled:
                self.printEnabled = False
                self.pb_startLog.setEnabled(True)
                self.pb_stopLog.setEnabled(False)
                self.cb_logEnable.setEnabled(True)
            else:
                self.printEnabled = True
                self.pb_startLog.setEnabled(False)
                self.pb_stopLog.setEnabled(True)
                self.cb_logEnable.setEnabled(False)
        else:
            self.logMessage("WARNING", "If no port is opened the log cannot be enabled. "
                                   "Please open a COM before enabling the log.")

    def printTrace(self, trace):
        if self.printEnabled:
            self.logMessage("TRACE", trace)
            if self.cb_logEnable.isChecked():
                self.parseTraces(trace)

    def calibrationCommand(self):
        if self.com.isOpen():
            self.com.sendCommand(0)
        else:
            self.logMessage("WARNING", "Unable to calibrate ESC. No COM is opened!")

    def parseTraces(self, trace):
        col = 0
        cleanTrace = str(trace).replace(";\r\n", "", 1)
        traceColumns = cleanTrace.split(";")
        for i in traceColumns:
            value = i.split(":")
            if value[0] == "t":
                value[1] = int(value[1])
                col = 1
            if value[0] == "TH":
                value[1] = float(value[1])
                col = 2
            if value[0] == "WS":
                volts = (float(value[1])*5)/1024
                value[1] = X6*math.pow(volts, 6)+X5*math.pow(volts, 5)+X4*math.pow(volts, 4)+\
                           X3*math.pow(volts, 3)+X2*math.pow(volts, 2)+X1*volts+ORD
                col = 3
            if value[0] == "MS":
                value[1] = int(value[1])
                col = 4
            if value[0] == "P":
                col = 5
            self.sheet.cell(self.rows, col, value[1])
        self.rows += 1

    def disableAuto(self):
        self.sldr_finalSpeedValue.setEnabled(False)
        self.lbl_finalSpeedValue.setEnabled(False)
        self.lbl_finalSpeedValueVar.setEnabled(False)
        self.lbl_percentAuto.setEnabled(False)
        self.lbl_timeElapsed.setEnabled(False)
        self.txt_timeElapsed.setEnabled(False)
        self.prog_automaticControl.setEnabled(False)
        self.pb_timeElapsed.setEnabled(False)
        self.lbl_ms.setEnabled(False)
        self.pb_startAutoControl.setEnabled(False)
        self.pb_stopAutoControl.setEnabled(False)

    def enableAuto(self):
        self.sldr_finalSpeedValue.setEnabled(True)
        self.lbl_finalSpeedValue.setEnabled(True)
        self.lbl_finalSpeedValueVar.setEnabled(True)
        self.lbl_percentAuto.setEnabled(True)
        self.lbl_timeElapsed.setEnabled(True)
        self.txt_timeElapsed.setEnabled(True)
        self.prog_automaticControl.setEnabled(True)
        self.pb_timeElapsed.setEnabled(True)
        self.lbl_ms.setEnabled(True)
        if not self.autoControlOnGoing:
            self.pb_startAutoControl.setEnabled(True)
            self.pb_stopAutoControl.setEnabled(False)

    def disableManual(self):
        self.speedKnob.setEnabled(False)
        self.lbl_currentSpeedValue.setEnabled(False)
        self.lbl_currentSpeedValueVar.setEnabled(False)
        self.lbl_percentManual.setEnabled(False)
        self.pb_lessSpeed.setEnabled(False)
        self.txt_incrementSpeedValue.setEnabled(False)
        self.pb_moreSpeed.setEnabled(False)
        self.lbl_InsertSpeedValue.setEnabled(False)
        self.txt_setSpeedValue.setEnabled(False)
        self.pb_setSpeedValue.setEnabled(False)
        self.lbl_acceleration.setEnabled(False)
        self.rb_accEnabled.setEnabled(False)
        self.rb_accDefault.setEnabled(False)
        self.pb_startManualControl.setEnabled(False)
        self.pb_stopManualControl.setEnabled(False)
        if self.rb_accEnabled.isChecked():
            self.sb_acc.setEnabled(False)

    def enableManual(self):
        self.speedKnob.setEnabled(True)
        self.lbl_currentSpeedValue.setEnabled(True)
        self.lbl_currentSpeedValueVar.setEnabled(True)
        self.lbl_percentManual.setEnabled(True)
        self.pb_lessSpeed.setEnabled(True)
        self.txt_incrementSpeedValue.setEnabled(True)
        self.pb_moreSpeed.setEnabled(True)
        self.lbl_InsertSpeedValue.setEnabled(True)
        self.txt_setSpeedValue.setEnabled(True)
        self.pb_setSpeedValue.setEnabled(True)
        self.lbl_acceleration.setEnabled(True)
        self.rb_accEnabled.setEnabled(True)
        self.rb_accDefault.setEnabled(True)
        if not self.manualControlOnGoing:
            self.pb_startManualControl.setEnabled(True)
            self.pb_stopManualControl.setEnabled(False)
        if self.rb_accEnabled.isChecked():
            self.sb_acc.setEnabled(True)

    def switchAutoManual(self):
        if self.rb_manualControl.isChecked():
            self.autoControlOnGoing = False
            self.disableAuto()
            self.sldr_finalSpeedValue.setValue(0)
            self.enableManual()
        elif self.rb_autoControl.isChecked():
            self.manualControlOnGoing = False
            self.disableManual()
            self.speedKnob.setValue(0)
            self.enableAuto()

    def activateControl(self):
        if self.rb_manualControl.isChecked():
            if self.com.isOpen():
                if not self.manualControlOnGoing:
                    self.manualControlOnGoing = True
                    self.pb_stopManualControl.setEnabled(True)
                    self.pb_startManualControl.setEnabled(False)
                    self.com.sendCommand(int(self.lbl_currentSpeedValueVar.text()) + 100)
                else:
                    self.manualControlOnGoing = False
                    self.pb_stopManualControl.setEnabled(False)
                    self.pb_startManualControl.setEnabled(True)
                    self.speedKnob.setValue(0)
                    self.com.sendCommand(290)
                    self.com.sendCommand(100)
            else:
                self.logMessage("WARNING", "No serial COM is opened. Manual control could not be activated!")
        if self.rb_autoControl.isChecked():
            if self.com.isOpen():
                if not self.autoControlOnGoing:
                    self.autoControlOnGoing = True
                    self.pb_startAutoControl.setEnabled(False)
                    self.pb_stopAutoControl.setEnabled(True)
                else:
                    self.autoControlOnGoing = False
                    self.pb_startAutoControl.setEnabled(True)
                    self.pb_stopAutoControl.setEnabled(False)
            else:
                self.logMessage("WARNING", "No serial COM is opened. Automatic control could not be activated!")

    def setAcc(self):
        self.com.sendCommand(int(self.sb_acc.text()) + 201)

    def disableAcc(self):
        if self.rb_accDefault.isChecked():
            if self.com.isOpen():
                self.com.sendCommand(290);
            self.sb_acc.setEnabled(False)
        elif self.rb_accEnabled.isChecked():
            if self.com.isOpen():
                self.com.sendCommand(int(self.sb_acc.text()) + 201)
            self.sb_acc.setEnabled(True)

    def setSpeedValueManual(self):
        try:
            value = int(self.txt_setSpeedValue.text())
            if 0 <= value <= 100:
                self.speedKnob.setValue(value)
                self.lbl_currentSpeedValueVar.setText(str(value))
                self.txt_setSpeedValue.clear()
                if self.com.isOpen():
                    self.com.sendCommand(value+100)
        except:
            if self.txt_setSpeedValue.text() == "":
                self.logMessage("WARNING", "NaN. Please, input a valid number!")


    def increaseSpeedValue(self):
        try:
            previousValue = int(self.lbl_currentSpeedValueVar.text())
            increment = int(self.txt_incrementSpeedValue.text())
            nextValue = previousValue + int(increment)
            if nextValue <= 100:
                self.lbl_currentSpeedValueVar.setText(str(nextValue))
                self.speedKnob.setValue(nextValue)
                if self.manualControlOnGoing:
                    if self.com.isOpen():
                        self.com.sendCommand(nextValue + 100)
            else:
                self.lbl_currentSpeedValueVar.setText(str(100))
                self.speedKnob.setValue(100)
                if self.manualControlOnGoing:
                    if self.com.isOpen():
                        self.com.sendCommand(200)
        except:
            self.logMessage("WARNING", "NaN. Please, input a valid number!")

    def decreaseSpeedValue(self):
        try:
            previousValue = int(self.lbl_currentSpeedValueVar.text())
            decrement = int(self.txt_incrementSpeedValue.text())
            nextValue = previousValue - int(decrement)
            if nextValue >= 0:
                self.lbl_currentSpeedValueVar.setText(str(nextValue))
                self.speedKnob.setValue(nextValue)
                if self.manualControlOnGoing:
                    if self.com.isOpen():
                        self.com.sendCommand(nextValue+100)
            else:
                self.lbl_currentSpeedValueVar.setText(str(0))
                self.speedKnob.setValue(0)
                if self.manualControlOnGoing:
                    if self.com.isOpen():
                        self.com.sendCommand(100)
        except:
            if self.txt_incrementSpeedValue.text() == "":
                self.logMessage("WARNING", "NaN. Please, input a valid number!")

    def manualControlKnob(self, value):
        self.lbl_currentSpeedValueVar.setText(str(value))
        if self.manualControlOnGoing:
            if self.com.isOpen():
                self.com.sendCommand(value+100)

    def autoControlSldr(self, value):
        self.lbl_finalSpeedValueVar.setText(str(value))

    def showSelectedTasks(self):
        try:
            unsortedItems = []
            for x in self.list_availableCommands.selectedItems():
                movingTask = str(x.text())
                self.com.sendCommand(self.availableTasks.index(movingTask))
                self.list_availableCommands.takeItem(self.list_availableCommands.row(x))
                unsortedItems.append(movingTask)
            numOfItems = self.list_shownCommands.count()
            while numOfItems > 0:
                item = self.list_shownCommands.item(numOfItems-1).text()
                unsortedItems.append(str(item))
                self.list_shownCommands.takeItem(numOfItems-1)
                numOfItems = numOfItems-1
            sortedItems = self.arraySorting(unsortedItems)
            for j in range(len(sortedItems)):
                self.list_shownCommands.addItem(sortedItems[j])
        except:
            self.logMessage("WARNING", "No COM port has been selected yet! Go to connections and choose one.")

    def eraseSelectedTasks(self):
        unsortedItems = []
        for x in self.list_shownCommands.selectedItems():
            movingTask = str(x.text())
            self.com.sendCommand(self.availableTasks.index(movingTask))
            self.list_shownCommands.takeItem(self.list_shownCommands.row(x))
            unsortedItems.append(movingTask)
        numOfItems = self.list_availableCommands.count()
        while numOfItems > 0:
            item = self.list_availableCommands.item(numOfItems - 1).text()
            unsortedItems.append(str(item))
            self.list_availableCommands.takeItem(numOfItems - 1)
            numOfItems = numOfItems - 1
        sortedItems = self.arraySorting(unsortedItems)
        for j in range(len(sortedItems)):
            self.list_availableCommands.addItem(sortedItems[j])

    def arraySorting(self, arr):
        iterations = len(arr)
        if iterations > 1:
            while iterations > 1:
                for i in range(iterations-1):
                    task1 = arr[i]
                    task2 = arr[i+1]
                    taskValue1 = self.availableTasks.index(arr[i])
                    taskValue2 = self.availableTasks.index(arr[i+1])
                    if taskValue1 > taskValue2:
                        arr[i] = task2
                        arr[i+1] = task1
                iterations = iterations-1

        return arr

    def logMessage(self, priority, message):
        EofL = "\n"
        if priority == "INFO":
            self.txt_log.setTextColor(QColor("cyan"))
            self.txt_log.insertPlainText("INFO: ")
        if priority == "WARNING":
            self.txt_log.setTextColor(QColor("red"))
            self.txt_log.insertPlainText("WARNING: ")
        if priority == "TRACE":
            self.txt_log.setTextColor(QColor("white"))
            EofL = ""
        self.txt_log.insertPlainText(message + EofL)
        self.txt_log.moveCursor(QTextCursor.End)

    def loggingEnable(self):
        if self.cb_logEnable.isChecked():
            self.path = str(QFileDialog.getSaveFileName(self, "Save log file"))
            if self.path is not "":
                splitPath = self.path.split("/")
                name = splitPath[-1]
                try:
                    if name.split(".")[1] == "xlsx":
                        pass
                    else:
                        splitPath[-1] = name.split(".")[0] + ".xlsx"
                except:
                    splitPath[-1] = name + ".xlsx"
                self.path = '/'.join(splitPath)
                self.wb = Workbook()
                self.sheet = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
                self.sheet.title = "Raw data"
                self.sheet["A1"] = "TIME"
                for i in range(1, len(self.availableTasks)-1):
                    self.sheet.cell(self.rows, i+1, str(self.availableTasks[i]))
                self.rows += 1
                self.logMessage("INFO", "Log enabled successfully! "
                                        "To save the log remember to uncheck the log data "
                                        "feature whenever you want to finish capturing data.")
            else:
                self.logMessage("WARNING", "Could not create log file. No log will be saved.")
                self.cb_logEnable.click()
        else:
            if self.path is not "":
                self.logMessage("INFO", "Saving data... Depending on the file's size "
                                        "this process can take several minutes.")
                self.txt_log.repaint()
                self.wb.save(self.path)
                self.rows = 1
                self.logMessage("INFO", "Log saved successfully on path: " + self.path)
                self.txt_log.moveCursor(QTextCursor.End)

    def versionInfo(self):
        version = str(softwareVersionMajor) + "." + str(softwareVersionMinor) + "." + str(softwareVersionPatch)
        vMsg = "SW Version: v." + version + "\n"
        devInfo = "Developers: " + developers[0] + "\n"
        rights = "The distribution of this SW is not allowed by the owner.\n"
        moRights = "Rights exclusively granted to UPC Venturi."

        QMessageBox.information(self, "Version Info", vMsg + devInfo + "\n" + rights + moRights, QMessageBox.Close)


class traceThread(QThread):
    def __init__(self):
        QThread.__init__(self)

    def __del__(self):
        self.wait()

    def run(self):
        while ser.isOpen():
            trace = ""
            while True:
                byte = ser.read(1)
                trace += byte
                if byte == "\n":
                    break
            if str(trace) != "":
                self.emit(SIGNAL('Serialtrace(QString)'), str(trace))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.processEvents()
    main = MainWindow()
    main.resize(1024, 600)
    main.show()
    if executableGeneration:
        app.setWindowIcon(QIcon(resource_path("logoventuri.png")))
    else:
        app.setWindowIcon(QIcon(os.path.join("imgV", "logoventuri.png")))
    sys.exit(app.exec_())

